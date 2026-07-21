"""
Export Excel multi-onglets — propre, lisible et prêt pour Power BI.

Objectif : un seul classeur .xlsx qui RÉPOND À L'ÉTUDE, avec des onglets nommés
en clair (pas de noms de variables bruts) — un onglet par table métrique
(10 tables, une table = un sujet complet), organisés par axe :
    1. Vue d'ensemble (synthèse, bilan cas par cas)
    2. Couverture des listes d'arrêts (les deux univers)
    3. Niveau de provisionnement (le taux de chute + la distribution des écarts)
    4. Suivi des consignes (détail + tableau de bord par type de compte)
    5. Investigation orphelins (six angles)
    6. Fiabilité (recoupements inter-onglets)
    7. Modèle de données (dimension du run — pivot des relations Power BI)

Chaque onglet est une VRAIE table Excel (ListObject) → Power BI la détecte comme
table nommée (« Get Data → Excel » propose directement chaque table), avec :
    - en-tête figé + filtre automatique (table Excel) ;
    - formats nombre € / % / entier appliqués (Power BI lit la valeur sous-jacente,
      le format reste cosmétique → aucun risque de parsing) ;
    - largeurs de colonnes ajustées.
Un onglet « Sommaire » liste les tables et l'axe d'étude auquel chacune répond.

Données tidy uniquement (1 ligne d'en-tête + lignes de données, pas de cellules
fusionnées) : c'est la forme que Power BI charge le plus proprement.
"""

from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook import Workbook

# Plan des onglets : (clé métrique, nom onglet, axe d'étude, contenu).
# L'ordre = l'ordre d'apparition dans le classeur. Toute table absente du plan
# est ajoutée à la fin (sans description) ; toute clé du plan absente des tables
# est simplement ignorée.
ONGLETS: Tuple[Tuple[str, str, str, str], ...] = (
    ("synthese",             "Synthèse",                "Vue d'ensemble",
     "Tous les indicateurs de tête du run (une ligne, historisable)."),
    ("bilan_cas",            "Bilan cas par cas",       "Vue d'ensemble",
     "La réconciliation cas par cas : retrouvés / non retrouvés / encore au compte."),
    ("couverture",           "Couverture",              "Couverture des listes",
     "Les deux univers (colonne UNIVERS) : justification du compte / part de la revue MRM retrouvée."),
    ("chute",                "Taux de chute",           "Niveau de provisionnement",
     "Le taux de chute et ses ventilations métier (colonne AXE) : Ensemble, type de compte, ancienneté — × exercice."),
    ("distribution_ecarts",  "Distribution des écarts", "Niveau de provisionnement",
     "La distribution des écarts de PM par tranche de seuils (histogramme) — × exercice, avec ligne Ensemble de référence."),
    ("consignes",            "Consignes",               "Suivi des consignes",
     "Conformité + PM + taux de chute par consigne, les deux exercices (colonne EXERCICE)."),
    ("consignes_par_type_compte", "Consignes par type", "Suivi des consignes",
     "Tableau de bord : suivi des consignes par type de compte × consigne."),
    ("orphelins",            "Orphelins",               "Investigation orphelins",
     "Les orphelins compte sous six angles (colonne AXE) : type de compte, garantie, ancienneté, mois, clause, clés nulles."),
    ("controles_coherence",  "Contrôles cohérence",     "Fiabilité",
     "Recoupements inter-onglets : une grandeur = une valeur partout (attendu/obtenu)."),
    ("dim_run",              "Dimension du run",        "Modèle de données",
     "Le pivot du modèle en étoile : une ligne par run (CLE_RUN), reliée en 1-n à chaque table."),
)

# Palette AXA pour l'en-tête du Sommaire.
_AXA_NAVY  = "00008F"
_GRIS_DOUX = "F2F2F4"

_INVALID_SHEET = set(r':\/?*[]')


def _safe_sheet_name(name: str, used: set) -> str:
    """Nom d'onglet valide Excel : caractères interdits retirés, ≤ 31, unique."""
    clean = "".join(c for c in name if c not in _INVALID_SHEET).strip()[:31] or "Onglet"
    base, i = clean, 2
    while clean in used:
        suffix = f" {i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _number_format(col: str, series: pd.Series) -> Optional[str]:
    """Format nombre par nom de colonne (€ / % / entier), None si non numérique."""
    if series.dtype == bool or not pd.api.types.is_numeric_dtype(series):
        return None
    up = col.upper()
    if "TAUX" in up or "PCT" in up or "POIDS" in up:
        return '0.0"%"'                      # valeur déjà en points de %, ex. 45.3 → 45,3%
    if "PM" in up or "ECART" in up:
        return '#,##0 €'
    return '#,##0'                            # nb dossiers, identifiants, multiplicités


def _style_data_sheet(ws, df: pd.DataFrame, table_name: str) -> None:
    """En-tête figé + table Excel + formats nombre + largeurs sur une feuille déjà
    remplie par pandas (en-tête en ligne 1, données à partir de la ligne 2)."""
    n_rows, n_cols = len(df), len(df.columns)
    ws.freeze_panes = "A2"

    # Table Excel (ListObject) : filtre + bandes + nom détecté par Power BI.
    # Nécessite au moins une ligne de données ; sinon on se contente de l'en-tête.
    if n_rows >= 1 and n_cols >= 1:
        ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
            showFirstColumn=False, showLastColumn=False,
        )
        ws.add_table(table)

    # Formats nombre (cosmétique : Power BI lit la valeur brute).
    for j, col in enumerate(df.columns, start=1):
        fmt = _number_format(col, df[col])
        if fmt:
            for i in range(2, n_rows + 2):
                ws.cell(row=i, column=j).number_format = fmt
        # Largeur = max(en-tête, contenu), bornée.
        longest = max([len(str(col))] + [len(str(v)) for v in df[col].head(200)], default=10)
        ws.column_dimensions[get_column_letter(j)].width = min(max(longest + 2, 10), 45)


def _write_sommaire(book: Workbook, plan: List[Tuple[str, str, str, str]],
                    client: str, perimetre: str) -> None:
    """Onglet « Sommaire » en tête : titre + table (Onglet / Axe / Contenu)."""
    ws = book.create_sheet("Sommaire", 0)
    ws["A1"] = f"Restitution ITIP-FIAB — {client} ({perimetre})"
    ws["A1"].font = Font(bold=True, size=14, color=_AXA_NAVY)
    ws["A2"] = ("Back-test CORECO (compte réellement provisionné) vs MRM (estimation) — "
                "onglets prêts pour Power BI")
    ws["A2"].font = Font(italic=True, size=10)

    header_row = 4
    headers = ["Onglet", "Axe de l'étude", "Contenu du tableau"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_AXA_NAVY)
        c.alignment = Alignment(vertical="center")
    for k, (_, sheet, axe, desc) in enumerate(plan, start=header_row + 1):
        ws.cell(row=k, column=1, value=sheet)
        ws.cell(row=k, column=2, value=axe)
        ws.cell(row=k, column=3, value=desc)
        if (k - header_row) % 2 == 0:
            for j in range(1, 4):
                ws.cell(row=k, column=j).fill = PatternFill("solid", fgColor=_GRIS_DOUX)

    ws.freeze_panes = f"A{header_row + 1}"
    for j, w in enumerate((26, 26, 78), start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def export_excel(
    tables   : Dict[str, pd.DataFrame],
    path     : str,
    client   : str,
    perimetre: str,
) -> str:
    """
    Écrit toutes les tables métriques dans un classeur Excel propre (1 onglet par
    table, ordonné par axe d'étude) + un onglet Sommaire. Renvoie le chemin.

    Args:
        tables    : {nom métrique → DataFrame pandas} (cf. metrics.toutes_metriques).
        path      : chemin .xlsx de sortie (local / /dbfs/...).
        client    : libellé client (CLIENT_NAME).
        perimetre : libellé périmètre (_PERIMETRE).
    """
    # Ordre du plan, puis tables hors plan (sécurité : rien n'est perdu).
    known   = {k for k, *_ in ONGLETS}
    ordered = [(k, s, a, d) for (k, s, a, d) in ONGLETS if k in tables]
    extras  = [(k, k, "Autre", "") for k in tables if k not in known]
    plan    = ordered + extras

    used: set = set()
    plan = [(k, _safe_sheet_name(s, used), a, d) for (k, s, a, d) in plan]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for key, sheet, _axe, _desc in plan:
            df = tables[key]
            df.to_excel(writer, sheet_name=sheet, index=False)
            _style_data_sheet(writer.sheets[sheet], df, table_name=f"T_{key}")
        _write_sommaire(writer.book, plan, client, perimetre)

    return path
